import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ================= 1. 定义完整数据（5大阶段，7列信息） =================
headers = [
    "项目阶段",
    "阶段目的（为什么要做）",
    "包含的具体步骤（严格按顺序）",
    "跑几轮Flow？（每轮的具体目的）",
    "输入文件",
    "输出文件 / 数据库（含Checkpoint）",
    "通关文牒（进入下一阶段的硬指标）"
]

# 数据行（严格按照上一轮最终版表格整理，无遗漏）
data_rows = [
    [
        "阶段1：架构与RTL可行性评估",
        "纸上算账。不动物理工具，只回答RTL写的面积/功耗在不在预算内，避免带必死架构进入后端。",
        "1. rtl_lint（语法检查）；2. power_est（活动因子预估功耗）；3. area_est（基于RTL结构预估面积）",
        "1 ~ 2 轮。第1轮：用默认估算模型跑出初步面积/功耗；第2轮（可选）：若超标，RTL架构师微调流水线/位宽后重跑验证。",
        "RTL代码（.v/.sv）；活动因子文件（.saif/.vcd）；工艺库估算模型（.lib/.db）",
        "面积预估报告（.area.est.rpt）；功耗预估报告（.power.est.rpt）。【不存Checkpoint】（无物理数据）",
        "预估面积 < 芯片预算面积；预估功耗 < 功耗预算的120%。超标则退回修改RTL。"
    ],
    [
        "阶段2：逻辑综合与DFT",
        "锻造零件+贴标签。把RTL变成具体门级网表，插入测试扫描链（DFT），生成后端物理实现的考纲（SDC）。",
        "1. synth（综合，RTL→门级网表）；2. formal（形式验证，比对网表与RTL）；3. dft（插入扫描链）；4. sdc_gen（生成时序约束）",
        "2 ~ 3 轮。第1轮：跑不带DFT的纯综合，确认网表面积和逻辑级数合理；第2轮：读入第1轮网表，插入扫描链并跑Formal；第3轮（可选）：若DFT覆盖率<95%，修改DFT脚本后重跑综合+DFT。",
        "RTL代码；标准单元库（.lib/.db）；DFT规则文件（.dft_spec）；综合约束（.synth.tcl）",
        "综合后网表（.v）；SDC约束（.sdc）；Formal通过报告（.formal.pass）；DFT覆盖率报告（.dft.rpt）。【必存Checkpoint_Netlist】（synth+formal通过后触发，保存netlist/目录）",
        "Formal验证100% Pass；DFT测试覆盖率 > 95%。否则退回修改RTL或DFT脚本。"
    ],
    [
        "阶段3：物理原型评估",
        "插旗探雷。只做快速摆放，不走时钟树和布线。用最短时间回答这块地皮会不会导致内部堵死？",
        "1. floorplan（定地皮、摆SRAM/PLL）；2. tapcell（撒阱接触单元）；3. gplace（全局布局，粗放标准单元）。【强制跳过】pdn、dplace、cts、route。",
        "1 ~ 2 轮。第1轮：采用CBR推荐的保守Floorplan（Macro间距放大10%）快速跑；第2轮（仅当第1轮拥塞爆红时）：微调Macro间距/旋转方向后重跑验证。",
        "阶段2输出的网表（.v）和SDC（.sdc）；物理库（.lef）；IO引脚位置文件（.io）",
        "拥塞地图（.congestion.rpt）；密度分布（.density.rpt）；未布线布局库（.enc/.odb）。【必存Checkpoint_Placement】（gplace完成后触发，保存placement.db/）",
        "拥塞溢出率（Overflow）< 1%且无深红拥堵区。爆红则退回修改Floorplan，不进入下一阶段。"
    ],
    [
        "阶段4：时序与物理收敛（最磨人阶段）",
        "盖楼通水电+极限拉扯。铺电网、精细摆单元、建时钟树、布金属线。Setup和Hold是死对头，必须反复博弈逼近平衡。",
        "1. pdn（铺电源/地线网格）；2. dplace（详细布局，消除重叠）；3. resize（换大单元修时序）；4. cts（时钟树综合）；5. groute（全局布线规划）；6. droute（详细布线，铺真实金属线）；7. post-route opt（布线后修Hold违例）",
        "5 ~ 10 轮。第1~2轮（基线）：跑通完整CTS+布线，获取初始时序基线（TNS最负，不要慌）；第3~5轮（修Setup）：逐轮调整时钟树缓冲器尺寸、开关面积恢复，观察WNS向0靠近；第6~8轮（修Hold）：Setup清零后插Delay Buffer修Hold（注意：插Buffer会恶化Setup，需交替进行）；第9~10轮（微调）：仅剩零星违例，手动锁定部分单元，做增量ECO冲刺清零。",
        "阶段3输出的布局库；寄生提取规则（.tluplus/.itf）；完整标准单元库（.lib）；更新的SDC（如有ECO改动）",
        "时序报告（WNS/TNS，.timing.rpt）；布通率报告（.route.rpt）；带时钟走线的完整库（.enc/.odb）。【必存Checkpoint_CTS】（cts完成后触发，保存cts.db/）；【必存Checkpoint_Route】（droute完成后，post-route opt之前触发，保存route.db/）",
        "Setup WNS > 0 且 Hold WNS > 0；同时布线溢出率 < 0.5%。否则调整CTS参数或微调Floorplan，回退到对应Checkpoint重跑。"
    ],
    [
        "阶段5：签核与流片",
        "拿到代工厂的入场券。做最后物理补丁，通过物理验证（DRC/LVS）和全工艺角（Corner）时序签核，导出芯片底片（GDS）。",
        "1. filler（填充空白，保证阱连续）；2. metal_fill（加金属密度填充，满足CMP打磨工艺）；3. write_gds（导出GDSII）。【并行执行】全Corner STA + DRC/LVS + IR-Drop分析。",
        "2 ~ 4 轮。第1轮（扫描）：跑完整DRC/LVS，必然会爆出几十到几百个违例；第2轮（修复）：在GUI里手动拉宽间距、插天线二极管，修复简单违例后重跑；第3~4轮（清零）：少数顽固违例需反复绕线（eco_route），每修一轮必须重跑全芯片DRC验证，直到0 Error。",
        "阶段4输出的最终收敛库；代工厂签核规则文件（DRC/LVS Rule Deck）；签核级STA库（PrimeTime专用）；IR-Drop分析模型",
        "GDSII文件（.gds）；DRC/LVS 0 Error报告（.drc.rpt/.lvs.rpt）；IR-Drop报告（.ir.rpt）；全Corner STA签核报告（.signoff.rpt）。【必存Checkpoint_Golden】（write_gds成功且DRC/LVS清零后触发，将gds/+脚本版本号永久打包归档）",
        "DRC违例=0；LVS比对=完全匹配（Pass）；IR-Drop（压降）< 3% VDD；所有工艺角（Corner）时序WNS > 0。任一不满足，进行小范围ECO微调后重跑此阶段。"
    ]
]

# ================= 2. 创建 Excel 工作簿并写入数据 =================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Full_Flow"

# 写入表头（加粗、蓝色背景）
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 写入数据行（左对齐，垂直居中，自动换行）
for row_idx, row_data in enumerate(data_rows, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ================= 3. 设置列宽（确保所有文字清晰可见） =================
col_widths = {
    'A': 20,  # 阶段名称
    'B': 28,  # 阶段目的
    'C': 32,  # 具体步骤
    'D': 42,  # 轮次详解（内容最多，给最宽）
    'E': 28,  # 输入文件
    'F': 36,  # 输出文件（含Checkpoint）
    'G': 34   # 通关文牒
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# 冻结首行（方便滚动查看）
ws.freeze_panes = 'A2'

# ================= 4. 保存文件 =================
file_name = "chip_full_flow.xlsx"
wb.save(file_name)
print(f"✅ Excel 文件已成功生成：{file_name}")
print(f"📊 共包含 {len(data_rows)} 个阶段（含阶段4的完整7个步骤及所有Checkpoint）。")